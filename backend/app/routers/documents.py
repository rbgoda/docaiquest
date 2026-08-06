from __future__ import annotations

import logging
import secrets

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import storage
from app.config import get_settings
from app.db import get_current_tenant, get_session
from app.models.documents import Document
from app.queue import enqueue_ingest, enqueue_rematch
from app.repositories import documents as repo
from app.security import CurrentUser, get_current_user, require_role
from app.services import documents as docs_service
# Re-exports — kept so any caller still using the underscore-prefixed
# helper names continues to work (TODO #25 conservative extraction).
_link_doc_to_requirement = docs_service.link_doc_to_requirement
_human_size = docs_service.human_size

from app.routers.documents_review import _build_review_metadata, ReviewPayload, BulkReviewPayload  # noqa: E402,F401 — re-exported for backward compat

router = APIRouter()

# ---- Self-learning classification -------------------------------------------
@router.post("/reclassify-other")
def reclassify_other(db: Session = Depends(get_session),
                     user: CurrentUser = Depends(get_current_user)) -> dict:
    """M46 · re-type the caller's 'other'/unclassified docs by reconciling each
    from its own AI summary (auto-assign), and learn the types. Documents
    product only; owner-scoped."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.services.type_reconciler import reconcile_doc
    docs = repo.list_unclassified(db)[:100]  # cap per pass
    changes: list[dict] = []
    for doc in docs:
        old = doc.doc_type or "other"
        try:
            new = reconcile_doc(db, doc)
        except Exception:  # noqa: BLE001
            new = None
        if new and new != old:
            changes.append({"docId": doc.id_external, "name": doc.name, "from": old, "to": new})
    return {"scanned": len(docs), "reclassified": len(changes), "changes": changes}


# ---- Supported formats ------------------------------------------------------
@router.get("/supported-types")
def supported_types() -> dict:
    """The upload formats the pipeline accepts — drives the file picker's accept
    filter + the 'Supported formats' hint. Single source of truth shared with
    storage.validate_upload (which hard-rejects anything else with a 415)."""
    from app.supported_types import SUPPORTED_UPLOAD_TYPES, accept_attr, UNSUPPORTED_NOTE
    return {"types": SUPPORTED_UPLOAD_TYPES, "accept": accept_attr(), "note": UNSUPPORTED_NOTE}


# §2 · declared BEFORE GET /{doc_id} so the static path isn't swallowed by the
# parametric doc-id route.
@router.get("/learned-types")
def learned_types(db: Session = Depends(get_session),
                  user: CurrentUser = Depends(get_current_user)) -> dict:
    """§2 · the caller's self-learned document-type vocabulary (slug, label,
    source ai|human, seen count, whether it's distilled). Documents product
    only; owner-scoped."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.repositories import learned_doc_types as ldt_repo
    return {"types": ldt_repo.list_all(db)}


class BulkPayload(BaseModel):
    action: str            # "delete" | "reclassify" | "share"
    docIds: list[str]
    groupId: int | None = None


@router.post("/bulk")
def bulk_action(payload: BulkPayload, db: Session = Depends(get_session),
                user: CurrentUser = Depends(get_current_user)) -> dict:
    """§A4 · bulk delete / reclassify / share-to-group over selected docs.
    Owner-scoped (each doc is resolved via owner-scoped get_row). Documents only."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.documents_scope import get_current_owner_user_pk
    from app.services import subscriptions as subs
    subs.enforce_feature(db, owner_user_id=get_current_owner_user_pk(), feature="bulk")  # M47 · Pro
    ids = payload.docIds or []
    done: list[str] = []
    if payload.action == "delete":
        for did in ids:
            if repo.delete_row(db, did) is not None:
                done.append(did)
        db.commit()
    elif payload.action == "reclassify":
        from app.services.type_reconciler import reconcile_doc
        for did in ids:
            doc = repo.get_row(db, did)
            if doc is not None and reconcile_doc(db, doc):
                done.append(did)
    elif payload.action == "share":
        if not payload.groupId:
            raise HTTPException(status_code=422, detail="groupId is required to share")
        from sqlalchemy import select as _sel
        from app.orm import DocumentGroupMember, DocumentGroupShare
        uid = get_current_owner_user_pk()
        member = db.scalar(_sel(DocumentGroupMember).where(
            DocumentGroupMember.group_id == payload.groupId,
            DocumentGroupMember.user_id == uid))
        if member is None:
            raise HTTPException(status_code=403, detail="Not a member of this group")
        tid = get_current_tenant()
        for did in ids:
            doc = repo.get_row(db, did)
            if doc is None or doc.owner_user_id != uid:
                continue
            exists = db.scalar(_sel(DocumentGroupShare).where(
                DocumentGroupShare.document_pk == doc.pk,
                DocumentGroupShare.group_id == payload.groupId))
            if exists is None:
                db.add(DocumentGroupShare(tenant_id=tid, document_pk=doc.pk, group_id=payload.groupId))
                done.append(did)
        db.commit()
    else:
        raise HTTPException(status_code=422, detail="Unknown action")
    return {"action": payload.action, "done": done, "count": len(done)}


@router.post("/workspace/sync")
async def workspace_sync(db: Session = Depends(get_session),
                         user: CurrentUser = Depends(get_current_user)) -> dict:
    """§5 · build the caller's encrypted workspace.sqlite and store it in THEIR
    Google Drive (`docaiq_docs/.workspace/`). Documents product only."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.documents_scope import get_current_owner_user_pk
    from app.services import subscriptions as subs
    from app.services import workspace_export
    uid = get_current_owner_user_pk()
    subs.enforce_feature(db, owner_user_id=uid, feature="workspace")  # M47 · Pro feature
    return await workspace_export.sync_to_drive(
        db, tenant_id=get_current_tenant(), owner_user_id=uid)


@router.get("/workspace/status")
def workspace_status(db: Session = Depends(get_session),
                     user: CurrentUser = Depends(get_current_user)) -> dict:
    """§5 · the caller's workspace-in-Drive sync status."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from sqlalchemy import select as _sel
    from app.orm import WorkspaceSync
    from app.documents_scope import get_current_owner_user_pk
    row = db.scalar(_sel(WorkspaceSync).where(
        WorkspaceSync.owner_user_id == get_current_owner_user_pk()))
    if row is None:
        return {"synced": False, "storageMode": get_settings().documents_storage_mode}
    return {"synced": True, "docCount": row.doc_count, "sizeBytes": row.size_bytes,
            "syncedAt": row.synced_at.isoformat() if row.synced_at else None,
            "storageMode": get_settings().documents_storage_mode}


@router.get("/learned-types/{slug}/candidates")
def learned_type_candidates(slug: str, db: Session = Depends(get_session),
                            user: CurrentUser = Depends(get_current_user)) -> dict:
    """§A3 · weakly-typed docs that look like learned type `slug` (centroid match)."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.services.type_reconciler import find_candidates
    return {"slug": slug, "candidates": find_candidates(db, slug)}


class ApplyTypePayload(BaseModel):
    docIds: list[str]


@router.post("/learned-types/{slug}/apply")
def learned_type_apply(slug: str, payload: ApplyTypePayload,
                       db: Session = Depends(get_session),
                       user: CurrentUser = Depends(get_current_user)) -> dict:
    """§A3 · apply learned type `slug` to the given weak docs (owner-scoped)."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.services.type_reconciler import apply_type_to_docs
    applied = apply_type_to_docs(db, slug, payload.docIds or [])
    return {"slug": slug, "applied": applied, "count": len(applied)}


@router.get("/search")
def search_documents(q: str = Query(..., min_length=1),
                     db: Session = Depends(get_session),
                     user: CurrentUser = Depends(get_current_user)) -> dict:
    """§A5 · metadata-first keyword search across the caller's own documents.

    Searches document name, doc type, extracted fields, and chunk text with
    Postgres full-text + trigram ranking. No embeddings, no semantic matching —
    only literal keyword matches. Returns the best matching chunk per doc so
    the UI can jump to it."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from app.documents_scope import get_current_owner_user_pk
    from app.services.document_search import keyword_search_documents

    uid = get_current_owner_user_pk()
    tid = get_current_tenant()

    rows = keyword_search_documents(db, q, tenant_id=tid, owner_user_id=uid)

    results = [{
        "docId": r["id_external"],
        "name": r["name"],
        "page": r["page"],
        "snippet": r["snippet"],
        "score": r["score"],
    } for r in rows]

    return {"query": q, "results": results}


@router.get("/extractions/export")
def export_extractions(format: str = Query("json", pattern="^(json|csv|xlsx)$"),
                       db: Session = Depends(get_session),
                       user: CurrentUser = Depends(get_current_user)):
    """§A6 · export the caller's extracted fields across all their docs as JSON,
    a flat CSV, or a real Excel (.xlsx) workbook. Documents product only; owner-scoped."""
    if get_settings().product != "documents":
        raise HTTPException(status_code=404, detail="Not available")
    from sqlalchemy import select as _sel
    from app.orm import Document as _Doc
    from app.documents_scope import get_current_owner_user_pk
    from app.services import subscriptions as subs
    uid = get_current_owner_user_pk()
    subs.enforce_feature(db, owner_user_id=uid, feature="export")  # M47 · Pro feature
    where = [_Doc.tenant_id == get_current_tenant(), _Doc.is_archived.is_(False)]
    if uid is not None:
        where.append(_Doc.owner_user_id == uid)
    docs = db.scalars(_sel(_Doc).where(*where).order_by(_Doc.pk)).all()

    def _flat(ef: dict | None) -> dict:
        f = (ef or {}).get("fields") if isinstance(ef, dict) and isinstance(ef.get("fields"), dict) else (ef or {})
        out = {}
        for k, v in (f or {}).items():
            if isinstance(v, (str, int, float)) or v is None:
                out[k] = v
        return out

    if format == "json":
        return {"extractions": [{"docId": d.id_external, "name": d.name,
                                 "docType": d.doc_type, "fields": _flat(d.extracted_fields)}
                                for d in docs]}
    # CSV / XLSX — union of all field keys → one row per document.
    import io as _io
    rows = [{"docId": d.id_external, "name": d.name, "docType": d.doc_type or "",
             **_flat(d.extracted_fields)} for d in docs]
    keys = ["docId", "name", "docType"]
    for r in rows:
        for k in r:
            if k not in keys:
                keys.append(k)

    if format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extractions"
        ws.append(keys)
        for r in rows:
            ws.append(["" if r.get(k) is None else r.get(k) for k in keys])
        bio = _io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            iter([bio.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=docaiq-extractions.xlsx"})

    import csv
    buf = _io.StringIO()
    w = csv.DictWriter(buf, fieldnames=keys, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow(r)
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=docaiq-extractions.csv"})


# ---- Reads ------------------------------------------------------------------
@router.get("", response_model=dict[str, Document])
def list_documents(
    include_archived: bool = Query(False, description="When true, include soft-archived documents in the result."),
    scope: str = Query("all", description="'personal' = only my own, non-group docs (Documents tab); 'all' = include group-shared docs."),
    db: Session = Depends(get_session),
) -> dict:
    return repo.list_all_map(db, include_archived=include_archived,
                             personal_only=(scope == "personal"))


# ── Manual document-type override (HITL) ─────────────────────────────────
# MUST be registered before /{doc_id} so "_meta" isn't captured as a doc_id.

@router.get("/_meta/doc-types")
def list_doc_types(_user: CurrentUser = Depends(get_current_user)) -> dict:
    """The classifier's known doc_type enum — powers the manual Type editor's
    autocomplete."""
    from app.agents.classifier import DOC_TYPES
    return {"docTypes": list(DOC_TYPES)}


@router.get("/{doc_id}", response_model=Document)
def get_document(doc_id: str, db: Session = Depends(get_session)) -> dict:
    row = repo.get(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    return row


@router.get("/{doc_id}/related")
def related_documents(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """Powers the 'Linked' tab. Returns:
    - duplicates: near-copies (same identifier / same issuer+amount+date) + any exact sha256 match.
      Exact byte-copies are normally blocked at upload; this catches re-scanned/re-exported copies
      with a different name or format.
    - related: documents that share a graph entity (same person/org via the reconciled canonical).
    Owner-scoped."""
    from app.services import related_docs
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    return {
        "docId": doc_id,
        "duplicates": related_docs.find_duplicates(db, doc),
        "related": related_docs.find_related(db, doc),
    }


@router.get("/{doc_id}/schema-json")
def schema_shaped_json(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """The document's extracted values rendered in its APPROVED schema's shape (fields in order,
    missing fields as null, each field's source marked). Falls back to the universal envelope when
    the type has no approved schema. Powers the JSON tab's 'Schema' view + export."""
    from app.routers.doc_chat import _reveal_fn
    from app.services import schema_json
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    result = schema_json.schema_shaped(db, doc)
    # Detokenize PII in the schema record for authorized viewers — same as the JSON /
    # Markdown exports. Without this the Schema view showed masked '[PERSON_1]' even
    # after the user revealed PII. _reveal_fn is a no-op when the doc isn't revealed.
    reveal = _reveal_fn(db, doc_id)

    def _reveal_deep(v):
        if isinstance(v, str):
            return reveal(v)
        if isinstance(v, list):
            return [_reveal_deep(x) for x in v]
        if isinstance(v, dict):
            return {k: _reveal_deep(x) for k, x in v.items()}
        return v
    if isinstance(result.get("record"), dict):
        result["record"] = _reveal_deep(result["record"])
    return {"docId": doc_id, **result}


@router.get("/{doc_id}/coverage")
def extraction_coverage(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """Deterministic per-doc extraction-coverage audit: how many salient page values
    (numbers/dates) the structured extraction captured, plus an explicit list of any it
    missed. Proves the lossless-chunk guarantee and grades structured coverage. No LLM."""
    from app.services import extraction_coverage as cov
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    return cov.coverage_report(db, doc)


class ChunkPatchPayload(BaseModel):
    text: str | None = None
    disabled: bool | None = None


@router.get("/{doc_id}/chunks")
def list_chunks(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """Chunk-inspection tab: the retrieval units for this doc, in reading order, with the source
    bbox + the disabled flag. These are exactly what BM25 + vector search see (disabled skipped)."""
    from app.orm import DocumentChunk
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    rows = db.scalars(select(DocumentChunk).where(DocumentChunk.document_pk == doc.pk)
                      .order_by(DocumentChunk.page, DocumentChunk.chunk_index)).all()
    return {"docId": doc_id, "chunks": [{
        "pk": c.pk, "index": c.chunk_index, "page": c.page, "kind": c.kind,
        "text": c.text, "bbox": c.bbox, "disabled": bool(c.disabled),
    } for c in rows]}


@router.patch("/{doc_id}/chunks/{chunk_pk}")
def patch_chunk(doc_id: str, chunk_pk: int, payload: ChunkPatchPayload,
                db: Session = Depends(get_session)) -> dict:
    """Edit a chunk's text (re-embeds it so vector search reflects the correction; tsv is a
    generated column so BM25 auto-updates) and/or toggle it in/out of retrieval."""
    from app.orm import DocumentChunk
    from app.embeddings import embed
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    c = db.scalar(select(DocumentChunk).where(
        DocumentChunk.pk == chunk_pk, DocumentChunk.document_pk == doc.pk))
    if c is None:
        raise HTTPException(status_code=404, detail="Chunk not found")
    if payload.text is not None and payload.text.strip() and payload.text != c.text:
        c.text = payload.text
        try:
            c.embedding = embed([payload.text])[0]
        except Exception:  # noqa: BLE001 — keep old vector if re-embed fails; text/tsv still fixed
            pass
    if payload.disabled is not None:
        c.disabled = bool(payload.disabled)
    db.commit()
    return {"pk": c.pk, "text": c.text, "disabled": bool(c.disabled)}


def _render_vision_markdown_bg(doc_pk: int) -> None:
    """Background task: render vision markdown for a document and cache it.
    Creates its own DB session so it doesn't depend on the request-scoped session."""
    from app.db import SessionLocal
    from app.orm import Document
    from app.services import markdown_export
    _bg_db = SessionLocal()
    try:
        doc = _bg_db.get(Document, doc_pk)
        if doc is None:
            return
        body = markdown_export.build_vision_markdown(_bg_db, doc)
        if body and body != markdown_export.build_full_markdown(_bg_db, doc):
            import datetime as _dt
            doc.rendered_markdown = body
            doc.rendered_markdown_at = _dt.datetime.now(_dt.timezone.utc)
            doc.rendered_markdown_model = get_settings().vision_qwen_model
            _bg_db.commit()
    except Exception:
        _bg_db.rollback()
    finally:
        _bg_db.close()


@router.get("/{doc_id}/markdown/full")
def export_full_markdown(
    doc_id: str,
    force: bool = False,
    raw: bool = False,
    db: Session = Depends(get_session),
) -> dict:
    """Whole-document Markdown — ALL users, no page cap, no LLM. Serves a human-corrected
    override if one exists (PUT below), else the deterministic build from the parsed text.
    Returns {docId, format, body, edited, blockMap?, annotatedBody?}.

    When ``rendered_markdown`` is not yet cached the endpoint returns immediately with
    the chunk-based render and schedules a background vision render to warm the cache
    for the next request — this keeps the Blocks/Rendered views responsive on first load.
    Use ``force=true`` to re-run vision synchronously.  Use ``raw=true`` to always get
    the chunk-based render even when a vision cache or override exists."""
    from app.orm import DocumentTextOverride
    from app.services import markdown_export

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    # Include block_map for frontend clickable-markdown sync when available
    block_map = getattr(doc, "block_map", None) or None
    # User-saved override takes precedence for the default view, but explicit
    # force/raw params bypass it so the Enhanced/Raw toggle still works.
    ovr = db.scalar(select(DocumentTextOverride).where(DocumentTextOverride.document_pk == doc.pk))
    if ovr is not None and not force and not raw:
        # User-saved override (always chunk-based) takes precedence for the
        # default view.  force/raw params bypass it so the Enhanced/Raw toggle
        # still fetches real vision / raw chunk content respectively.
        return {"docId": doc_id, "format": "markdown",
                "body": ovr.markdown, "edited": True,
                "annotatedBody": ovr.markdown if "<!-- block:" in ovr.markdown else None,
                "blockMap": block_map,
                "rendered": None}

    # ── Build the annotated body FIRST (always fast, no LLM) ──────────────
    # When block_map exists, generate the annotated markdown for per-block
    # editing + PDF sync in the Blocks view.  Returns null when no block_map.
    annotated_body = markdown_export.build_annotated_markdown(db, doc) if block_map else None

    # ── Build the formatted (readable) body ──────────────────────────────────
    # Prefer the cached vision-rendered markdown; fall back to the fast
    # deterministic chunk-based render.  When the cache is cold we skip the slow
    # vision path and schedule a background render so the next request is warm.
    # ``raw=true`` always returns chunk-based markdown (skips vision entirely).
    formatted_body: str | None = None
    _did_vision = False  # track whether we actually ran vision (for caching)
    _is_vision_body = False  # whether the returned body IS vision-rendered
    if raw:
        # Explicit raw request — always chunk-based, no vision cache check
        formatted_body = markdown_export.build_full_markdown(db, doc)
        _is_vision_body = False
    elif doc.rendered_markdown and not force:
        formatted_body = doc.rendered_markdown
        _is_vision_body = True
    elif force:
        # Explicit force-refresh — run vision synchronously (user asked).
        # Only cache the result if vision actually produced output (not a fallback).
        formatted_body = markdown_export.build_vision_markdown(db, doc)
        if formatted_body:
            _did_vision = True
            _is_vision_body = True
        else:
            formatted_body = markdown_export.build_full_markdown(db, doc)
            _is_vision_body = False
    elif not doc.rendered_markdown:
        # Cold cache: use fast chunk-based render; schedule vision for next time
        formatted_body = markdown_export.build_full_markdown(db, doc)
        _is_vision_body = False
        if formatted_body:
            import threading
            threading.Thread(target=_render_vision_markdown_bg, args=(doc.pk,), daemon=True).start()
    if not formatted_body:
        formatted_body = markdown_export.build_full_markdown(db, doc)
        _is_vision_body = False
    if not formatted_body:
        raise HTTPException(status_code=409,
                            detail="No parsed text is available for this document yet")
    # Cache vision-rendered markdown when we actually ran vision (not chunk-based)
    if _did_vision and (not doc.rendered_markdown or force):
        try:
            import datetime as _dt
            doc.rendered_markdown = formatted_body
            doc.rendered_markdown_at = _dt.datetime.now(_dt.timezone.utc)
            doc.rendered_markdown_model = get_settings().vision_qwen_model
            db.commit()
        except Exception:  # noqa: BLE001 — caching is best-effort
            db.rollback()

    return {"docId": doc_id, "format": "markdown",
            "body": formatted_body,
            "annotatedBody": annotated_body,
            "edited": False,
            "blockMap": block_map,
            "rendered": "vision" if _is_vision_body else None,
            "_debug": {"raw": raw, "force": force,
                       "has_vision_cache": bool(doc.rendered_markdown),
                       "is_vision_body": _is_vision_body,
                       "body_len": len(formatted_body) if formatted_body else 0}}


class MarkdownOverridePayload(BaseModel):
    markdown: str
    changed_block_ids: list[str] | None = None


@router.put("/{doc_id}/markdown")
def save_markdown_override(
    doc_id: str,
    payload: MarkdownOverridePayload,
    reprocess: bool = Query(False),
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Save a human-corrected full-document Markdown as an override. Audited via
    field_edits (field_path='__markdown__').

    When ``?reprocess=true``, deletes the old chunks, re-chunks the edited text,
    regenerates contextual summaries, re-embeds, and re-runs fact extraction so
    retrieval + extracted fields reflect the edited content."""
    from app.orm import DocumentTextOverride, FieldEdit
    from app.services import markdown_export

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    md = payload.markdown or ""
    if not md.strip():
        raise HTTPException(status_code=400, detail="Markdown cannot be empty")
    row = db.scalar(select(DocumentTextOverride).where(DocumentTextOverride.document_pk == doc.pk))
    prev = row.markdown if row is not None else (markdown_export.build_full_markdown(db, doc) or "")
    if row is None:
        db.add(DocumentTextOverride(tenant_id=user.org_id, document_pk=doc.pk,
                                    markdown=md, edited_by=user.email))
    else:
        row.markdown = md
        row.edited_by = user.email
    db.add(FieldEdit(tenant_id=user.org_id, document_pk=doc.pk, field_path="__markdown__",
                     original_value=prev[:4000], new_value=md[:4000],
                     edited_by=user.email, reason="full-text edit"))
    db.commit()

    response = {"docId": doc_id, "format": "markdown", "body": md, "edited": True}

    if not reprocess:
        return response

    # ── Reprocess: re-chunk → re-embed → re-extract ──────────────────────
    import re

    from sqlalchemy import delete as sqla_delete

    from app.agents import fact_extractor
    from app.contextual import embedding_input, generate_contexts
    from app.embeddings import embed as embed_v1, embed_v2
    from app.ingestion import Chunk, chunk_pages
    from app.orm import DocumentChunk

    settings = get_settings()
    log = logging.getLogger("docaiq.markdown_reprocess")

    try:
        # 1. Chunk — block-aware when block_map exists; flat chunk_pages otherwise.
        has_block_map = getattr(doc, "block_map", None) is not None
        if has_block_map:
            from app.ingestion import _chunk_annotated_markdown
            chunks: list[Chunk] = _chunk_annotated_markdown(
                md, target=settings.chunk_target_chars,
                overlap=settings.chunk_overlap_chars,
                nfkc=settings.chunk_nfkc_normalize,
            )
            clean = re.sub(r"<!-- block:b_\w+ -->", "", md).strip()
        else:
            clean = re.sub(r"<!-- block:b_\w+ -->", "", md).strip()
            chunks = chunk_pages([(1, clean)])
        if not chunks:
            log.warning("reprocess: doc pk=%s — chunk_pages produced zero chunks", doc.pk)
            return {**response, "reprocessed": False, "error": "No chunks produced from edited text"}

        chunk_texts = [c.text for c in chunks]

        # 2. Determine which chunks are affected (need fresh contexts).
        changed_set = set(payload.changed_block_ids) if payload.changed_block_ids else set()
        incremental = bool(changed_set and has_block_map)

        if incremental:
            # Save old context summaries before deletion (reuse for text-identical
            # unchanged chunks → skip the LLM context call).
            old_rows = db.scalars(
                select(DocumentChunk)
                .where(DocumentChunk.document_pk == doc.pk)
                .order_by(DocumentChunk.chunk_index)
            ).all()

            # Safety: if any old chunk has NULL block_ids (pre-existing data), fall
            # back to full reprocess — we can't tell which blocks they belong to.
            if old_rows and any(getattr(r, "block_ids", None) is None for r in old_rows):
                log.info("reprocess: doc pk=%s — old chunks have NULL block_ids, "
                         "falling back to full reprocess", doc.pk)
                incremental = False
            else:
                old_ctx_by_text: dict[str, str] = {}
                for r in old_rows:
                    if r.text and r.context_summary:
                        old_ctx_by_text[r.text] = r.context_summary

                affected = [
                    not chunk.block_ids or any(bid in changed_set for bid in chunk.block_ids)
                    for chunk in chunks
                ]
                log.info("reprocess: doc pk=%s — incremental: %s/%s chunks affected "
                         "(changed block IDs: %s)", doc.pk, sum(affected), len(chunks),
                         ",".join(sorted(changed_set)[:10]))

        if not incremental:
            affected = [True] * len(chunks)

        # 3. Delete old chunks (atomic — all or nothing with the commit below).
        deleted = db.execute(
            sqla_delete(DocumentChunk).where(DocumentChunk.document_pk == doc.pk)
        )
        db.flush()
        log.info("reprocess: doc pk=%s — deleted %s old chunks", doc.pk, deleted.rowcount)

        # 4. Generate contextual summaries — only for affected chunks.
        contexts: list[str] = [""] * len(chunks)
        to_gen = [(i, chunk_texts[i]) for i, a in enumerate(affected) if a]
        if to_gen:
            gen_texts = [t for _, t in to_gen]
            generated = generate_contexts(
                doc_text=clean,
                doc_name=doc.name or doc.id_external or "document",
                chunk_texts=gen_texts,
            )
            for (idx, _), ctx in zip(to_gen, generated):
                contexts[idx] = ctx or ""
            log.info("reprocess: doc pk=%s — generated %s contexts (LLM)", doc.pk, len(to_gen))

        # Reuse old context for text-identical unchanged chunks (zero LLM cost).
        if incremental:
            reused = 0
            for i in range(len(chunks)):
                if not affected[i] and chunk_texts[i] in old_ctx_by_text:
                    contexts[i] = old_ctx_by_text[chunk_texts[i]]
                    reused += 1
            if reused:
                log.info("reprocess: doc pk=%s — reused %s old contexts", doc.pk, reused)

        populated = sum(1 for c in contexts if c)
        log.info("reprocess: doc pk=%s — %s/%s contexts populated", doc.pk, populated, len(chunks))

        # 5. Embed v1 + v2 with contextual text (re-embed all — cheap, one batch call).
        contextual_texts = [
            embedding_input(chunk_texts[i], (contexts[i] if i < len(contexts) else None))
            for i in range(len(chunks))
        ]
        v1_vecs = embed_v1(contextual_texts)
        if settings.embed_v2_active:
            v2_vecs = embed_v2(contextual_texts)
        else:
            v2_vecs = None

        # 6. Create DocumentChunk rows (with embeddings already set).
        for i, chunk in enumerate(chunks):
            ctx = (contexts[i] if i < len(contexts) else None)
            db.add(DocumentChunk(
                tenant_id=user.org_id,
                document_pk=doc.pk,
                chunk_index=i,
                text=chunk.text,
                page=chunk.page,
                char_start=chunk.char_start,
                char_end=chunk.char_end,
                bbox=chunk.bbox,
                block_ids=list(chunk.block_ids) if chunk.block_ids else None,
                context_summary=(ctx or None),
                embedding=v1_vecs[i] if i < len(v1_vecs) else [],
                embedding_v2=(v2_vecs[i] if v2_vecs and i < len(v2_vecs) else None),
            ))

        db.flush()  # assign PKs so extraction can reference them
        response["chunksUpdated"] = sum(affected)
        response["chunksKept"] = len(chunks) - sum(affected)
        response["totalChunks"] = len(chunks)
        response["contextsPopulated"] = populated
        log.info("reprocess: doc pk=%s — %s chunks staged (%s updated, %s kept)",
                 doc.pk, len(chunks), sum(affected), len(chunks) - sum(affected))

        # 7. Re-run fact extraction (reuses existing doc_type; best-effort).
        classifier_doc_type = doc.doc_type or "universal"
        try:
            fx = fact_extractor.extract(
                db, document_pk=doc.pk, classifier_doc_type=classifier_doc_type,
            )
            if fx is not None:
                _old = doc.extracted_fields or {}
                _tl = _old.get("text_layer")
                doc.extracted_fields = fx.to_jsonb()
                if _tl:
                    doc.extracted_fields["text_layer"] = _tl
                response["fields"] = fx.fields
                response["docType"] = fx.schema_key
                response["confidence"] = fx.confidence
                log.info("reprocess: doc pk=%s — fields re-extracted (%s fields, conf=%.2f)",
                         doc.pk, len(fx.fields), fx.confidence)
        except Exception as e:
            log.warning("reprocess: doc pk=%s — fact extraction failed (best-effort): %s", doc.pk, e)

        # Commit chunks + fields atomically — no orphan chunks on extraction failure.
        db.commit()
        response["reprocessed"] = True

    except Exception as e:
        log.exception("reprocess: doc pk=%s — reprocess failed: %s", doc.pk, e)
        db.rollback()
        response["reprocessed"] = False
        response["error"] = str(e)

    return response


@router.delete("/{doc_id}/markdown/override")
def reset_markdown_override(
    doc_id: str,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Discard the human-corrected override and fall back to the deterministic build."""
    from app.orm import DocumentTextOverride
    from app.services import markdown_export

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    row = db.scalar(select(DocumentTextOverride).where(DocumentTextOverride.document_pk == doc.pk))
    if row is not None:
        db.delete(row)
        db.commit()
    block_map = getattr(doc, "block_map", None) or None
    return {"docId": doc_id, "format": "markdown",
            "body": markdown_export.build_full_markdown(db, doc), "edited": False,
            "blockMap": block_map,
            "annotatedBody": markdown_export.build_annotated_markdown(db, doc) if block_map else None}


# ── Translation & Export (M54) ───────────────────────────────────────────────

class TranslatePayload(BaseModel):
    target_language: str  # "fr", "de", "es", …


@router.post("/{doc_id}/translate")
def translate_document(doc_id: str, payload: TranslatePayload,
                       db: Session = Depends(get_session),
                       user: CurrentUser = Depends(get_current_user)) -> dict:
    """Translate the document's markdown into the target language while
    preserving ``<!-- block:b_XXXX -->`` markers. Returns the cached
    translation when available; calls the LLM on first request per language."""
    from app.services.document_translation import (
        SUPPORTED_LANGUAGES, translate_markdown,
    )

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    lang = (payload.target_language or "").strip().lower()
    if lang not in SUPPORTED_LANGUAGES:
        raise HTTPException(
            status_code=422,
            detail=f"Unsupported language '{lang}'. "
                   f"Supported: {', '.join(sorted(SUPPORTED_LANGUAGES.keys()))}",
        )

    try:
        entry = translate_markdown(
            db, doc, lang,
            tenant_id=get_current_tenant(),
            user_email=user.email if hasattr(user, "email") else None,
        )
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=str(e))

    return {
        "docId": doc_id,
        "language": lang,
        "body": entry["body"],
        "annotatedBody": entry.get("annotated_body", entry["body"]),
        "translatedAt": entry.get("translated_at"),
        "model": entry.get("model"),
        "cached": entry.get("cached", False),
        "source": entry.get("source", "chunks"),
        "truncated": entry.get("truncated", False),
        "warning": entry.get("warning"),
    }


@router.get("/{doc_id}/translations")
def list_translations(doc_id: str,
                      db: Session = Depends(get_session),
                      _user: CurrentUser = Depends(get_current_user)) -> dict:
    """Return all cached translations for a document, keyed by language code."""
    from app.services.document_translation import list_translations as _list

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    return {"docId": doc_id, "translations": _list(doc)}


@router.get("/{doc_id}/export")
def export_document(doc_id: str,
                    format: str = Query("json", pattern="^(json|csv)$"),
                    db: Session = Depends(get_session),
                    _user: CurrentUser = Depends(get_current_user)):
    """Export a single document's extracted fields as JSON or CSV.

    Returns a JSON response for ``format=json``; a streaming CSV download for
    ``format=csv`` (with ``Content-Disposition: attachment``)."""
    import csv as _csv
    import io as _io
    import re as _re

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    ef = getattr(doc, "extracted_fields", None) or {}
    # Normalise: extracted_fields is often {"fields": {...}, ...}
    fields = ef.get("fields", {}) if isinstance(ef, dict) and "fields" in ef else ef
    if not isinstance(fields, dict):
        fields = {}

    def _flat() -> dict[str, object]:
        out: dict[str, object] = {
            "docId": doc.id_external,
            "name": doc.name or "",
            "docType": doc.doc_type or "",
        }
        for k, v in fields.items():
            if isinstance(v, (str, int, float, bool)) or v is None:
                out[k] = v
        return out

    if format == "json":
        return {
            "docId": doc_id,
            "name": doc.name,
            "docType": doc.doc_type,
            "fields": fields,
        }

    # CSV
    row = _flat()
    keys = list(row.keys())
    buf = _io.StringIO()
    w = _csv.DictWriter(buf, fieldnames=keys, extrasaction="ignore")
    w.writeheader()
    w.writerow(row)
    buf.seek(0)
    from starlette.responses import StreamingResponse
    safe_name = _re.sub(r'[^a-zA-Z0-9._-]', '_', (doc.name or doc_id))[:64]
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={safe_name}-fields.csv",
        },
    )


@router.get("/{doc_id}/recall-gaps")
def recall_gaps(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """Structured-looking spans (dates, money, emails, IDs, …) in the parsed text that NO
    extracted field covers — candidate *missed* fields the reviewer can locate on the page
    and add (region→field). Each gap carries a located bbox where findable. Best-effort."""
    from app.services import recall_gap
    from app.agents.fact_extractor import _build_text_excerpt, _locate_field_bboxes

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    try:
        text, refs = _build_text_excerpt(db, doc.pk, full=True)
    except Exception:  # noqa: BLE001
        text, refs = "", []
    covered = recall_gap.collect_covered_values(doc.extracted_fields)
    gaps = recall_gap.find_gaps(text, covered, limit=20)
    try:
        gap_fields = {f"__gap_{i}": g["value"] for i, g in enumerate(gaps)}
        boxes = _locate_field_bboxes(db, doc.pk, gap_fields, refs)
        for i, g in enumerate(gaps):
            g["bbox"] = boxes.get(f"__gap_{i}")
    except Exception:  # noqa: BLE001
        pass
    return {"docId": doc_id, "gaps": gaps}


# ---- Streamed file download -------------------------------------------------
@router.get("/{doc_id}/file")
async def download_file(
    doc_id: str,
    db: Session = Depends(get_session),
    _user: CurrentUser = Depends(get_current_user),
) -> StreamingResponse:
    """Stream the underlying object from S3 through the backend. Bypasses
    CORS / presigned-URL host-mismatch issues for per-tenant deployments."""
    row = repo.get_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if not row.s3_key:
        # M46 · self-healing. A connector doc whose original was purged (s3_key
        # cleared) is transparently RE-PULLED from its source (Drive) on open,
        # so the viewer never dead-ends on a missing file. Owner scope is already
        # enforced by repo.get_row above.
        if (get_settings().product == "documents" and row.source == "drive"
                and row.source_ref):
            from app.connectors import drive as drive_mod
            from app.repositories import connectors as conn_repo
            acct = conn_repo.get(db, "drive")
            if acct is not None:
                import hashlib as _hashlib
                import io as _io
                try:
                    pulled = await drive_mod.get_backend().fetch(acct, row.source_ref)
                    # B7 · decrypt if this was a DocAIQ-encrypted Drive copy
                    # (no-op for plaintext / user-dropped files).
                    from app import drive_crypto
                    body = drive_crypto.decrypt_blob(row.owner_user_id, pulled.body)
                    tenant = get_current_tenant()
                    sha = row.sha256 or _hashlib.sha256(body).hexdigest()
                    key = f"{tenant}/documents/{sha[:2]}/{sha}-{secrets.token_hex(8)}"
                    storage.put_object(key, _io.BytesIO(body),
                                       content_type=pulled.content_type)
                    row.s3_key = key
                    db.commit()
                except Exception:  # noqa: BLE001
                    pass
    if not row.s3_key:
        raise HTTPException(
            status_code=404, detail="Document has no stored file (seeded demo doc)"
        )

    return StreamingResponse(
        storage.stream_object(row.s3_key),
        media_type=row.mime_type or "application/octet-stream",
        headers={"Content-Disposition": f'inline; filename="{storage.sanitize_filename(row.name)}"'},
    )


@router.get("/{doc_id}/sheets")
def document_sheets(
    doc_id: str,
    db: Session = Depends(get_session),
    _user: CurrentUser = Depends(get_current_user),
) -> dict:
    """Return an .xlsx/.xls workbook as JSON sheets so the viewer can render it
    as tables — PDF.js can't open spreadsheets. Owner-scoped via repo.get_row.
    Bounded (50 sheets × 500 rows × 30 cols) so a huge workbook can't blow up
    the response."""
    row = repo.get_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if not row.s3_key:
        raise HTTPException(status_code=404, detail="Document has no stored file")
    raw = storage.get_object_bytes(row.s3_key)
    if not raw:
        raise HTTPException(status_code=404, detail="Document file unavailable")
    import io as _io

    import openpyxl
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=415, detail=f"Not a readable spreadsheet: {e}")
    sheets = []
    for ws in wb.worksheets[:50]:
        rows = []
        for r in ws.iter_rows(max_row=500, max_col=30, values_only=True):
            rows.append(["" if c is None else str(c) for c in r])
        while rows and not any(c.strip() for c in rows[-1]):
            rows.pop()
        sheets.append({"name": ws.title, "rows": rows,
                       "truncated": (ws.max_row or 0) > 500 or (ws.max_column or 0) > 30})
    wb.close()
    return {"sheets": sheets}


# ---- Upload -----------------------------------------------------------------
@router.get("/{doc_id}/status")
def document_status(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """Lightweight polling endpoint for the frontend after upload."""
    from sqlalchemy import select, func
    from app.orm import DocumentChunk

    row = repo.get_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    chunk_count = db.scalar(
        select(func.count()).select_from(DocumentChunk).where(DocumentChunk.document_pk == row.pk)
    ) or 0
    return {
        "id": doc_id,
        "status": row.ingestion_status,   # None for seeded, else pending/processing/ready/failed
        "error": row.ingestion_error,
        "pages": row.pages,
        "chunks": chunk_count,
    }


@router.get("/{doc_id}/quality")
def document_quality(doc_id: str, db: Session = Depends(get_session)) -> dict:
    """M47 · Per-document indexing quality scorecard — chunk stats, embedding
    status, language detection, pipeline version. Displayed in the document
    detail view so users can verify their documents were processed correctly."""
    from sqlalchemy import select, func
    from app.orm import DocumentChunk

    row = repo.get_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    # Chunk stats
    chunks = db.execute(
        select(
            func.count().label("total"),
            func.count().filter(DocumentChunk.embedding_v2.isnot(None)).label("v2"),
            func.count().filter(DocumentChunk.pipeline_version >= 2).label("pv2"),
            func.avg(func.length(DocumentChunk.text)).label("avg_len"),
            func.min(func.length(DocumentChunk.text)).label("min_len"),
            func.max(func.length(DocumentChunk.text)).label("max_len"),
            func.max(DocumentChunk.pipeline_version).label("max_pv"),
        ).where(DocumentChunk.document_pk == row.pk)
    ).first()

    # Language detection on first 10 chunks
    lang_texts = db.execute(
        select(DocumentChunk.text).where(
            DocumentChunk.document_pk == row.pk
        ).order_by(DocumentChunk.chunk_index).limit(10)
    ).scalars().all()

    langs = {}
    mixed = False
    if lang_texts:
        try:
            from app.agents.indexing_critic import detect_languages, is_mixed_language
            langs = detect_languages(" ".join(t for t in lang_texts if t))
            mixed = is_mixed_language(langs)
        except Exception:
            pass

    return {
        "id": doc_id,
        "name": row.name,
        "type": row.doc_type,
        "status": row.ingestion_status,
        "pages": row.pages,
        "chunks": {
            "total": chunks.total if chunks else 0,
            "v2_embedded": chunks.v2 if chunks else 0,
            "v2_pipeline": chunks.pv2 if chunks else 0,
            "pipeline_version": chunks.max_pv if chunks else 0,
            "avg_len": round(chunks.avg_len or 0),
            "min_len": chunks.min_len or 0,
            "max_len": chunks.max_len or 0,
        },
        "language": {
            "detected": {k: round(v, 2) for k, v in sorted(langs.items(), key=lambda x: -x[1])[:4]},
            "mixed": mixed,
        },
        "embedding": {
            "v1_model": "all-MiniLM-L6-v2",
            "v1_dim": 384,
            "v2_model": "BAAI/bge-m3",
            "v2_dim": 1024,
        },
        # M47 · Per-field quality scores
        "field_quality": _compute_field_quality(row),
    }


def _compute_field_quality(doc) -> dict:
    """M47 · 6-dimensional per-field quality scoring.
    Weights: LLM confidence 25%, format validity 20%, bbox/position 15%,
    data consistency 15%, text clarity 15%, field uniqueness 10%."""
    ef = doc.extracted_fields or {}
    fields = ef.get("fields", {})
    bboxes = ef.get("field_bboxes", {})
    conf = ef.get("field_confidence", {})
    scores = {}

    for fname, fval in fields.items():
        if fname.startswith("_"): continue
        dims = {}
        # 1. LLM extraction confidence (25%)
        c = conf.get(fname)
        dims["llm_confidence"] = float(c) if isinstance(c, (int, float)) else 0.5
        # 2. Format validity (20%)
        dims["format_valid"] = _check_format(fname, fval)
        # 3. Bbox / positional validity (15%)
        dims["position_valid"] = 1.0 if fname in bboxes else 0.0
        # 4. Data consistency (15%) — cross-field validation
        dims["data_consistent"] = _check_consistency(fname, fval, fields)
        # 5. Text clarity (15%) — based on bbox and text_layer presence
        dims["text_clarity"] = _check_clarity(doc, fname, fval)
        # 6. Field uniqueness (10%) — value not duplicated across fields
        dims["field_unique"] = _check_uniqueness(fname, fval, fields)

        overall = round(
            dims["llm_confidence"] * 0.25
            + dims["format_valid"] * 0.20
            + dims["position_valid"] * 0.15
            + dims["data_consistent"] * 0.15
            + dims["text_clarity"] * 0.15
            + dims["field_unique"] * 0.10, 3
        )
        scores[fname] = {"overall": overall, "dimensions": dims}

    all_scores = [s["overall"] for s in scores.values()]
    return {
        "fields": scores,
        "avg_score": round(sum(all_scores) / len(all_scores), 3) if all_scores else 0,
        "low_scoring": [k for k, v in scores.items() if v["overall"] < 0.5],
        "missing_bboxes": [k for k in fields if k not in bboxes and not k.startswith("_")],
    }


def _check_consistency(fname: str, fval, all_fields: dict) -> float:
    """Cross-field consistency check. e.g. total = sum of line items, dates in order."""
    if not fval: return 0.5
    s = str(fval).strip()
    if not s: return 0.5
    # Invoice total vs line items: check that total > individual amounts
    if "total" in fname.lower() or "amount" in fname.lower() or "sum" in fname.lower():
        import re
        nums = re.findall(r"[\d,]+\.?\d*", s)
        if nums:
            return 0.8  # has a numeric value, plausible
        return 0.3  # total field should have numbers
    # Dates: start_date should be before end_date
    if "start" in fname.lower():
        end_key = fname.replace("start", "end")
        if end_key in all_fields:
            return 0.8  # both start and end exist
    if "end" in fname.lower() or "expiry" in fname.lower():
        return 0.8
    # Default: not a cross-checkable field
    return 0.7


def _check_clarity(doc, fname: str, fval) -> float:
    """Text clarity — is the field in a text-dense region (good OCR)?"""
    if not fval: return 0.5
    ef = doc.extracted_fields or {}
    bboxes = ef.get("field_bboxes", {})
    # Has bbox → text_locate confirmed it's findable
    if fname in bboxes: return 0.9
    # Check text_layer — if document has lots of words, it's well-OCR'd
    tl = ef.get("text_layer") or []
    if len(tl) > 100: return 0.7  # well-populated text layer
    if len(tl) > 0: return 0.5   # sparse but has some text
    return 0.3  # no text layer, likely scanned/image


def _check_uniqueness(fname: str, fval, all_fields: dict) -> float:
    """Field uniqueness — value should not be duplicated across fields."""
    if not fval: return 0.8
    s = str(fval).strip()
    if not s or len(s) < 3: return 0.8
    dupes = 0
    for k, v in all_fields.items():
        if k != fname and v and str(v).strip() == s:
            dupes += 1
    return 1.0 if dupes == 0 else max(0.3, 1.0 - dupes * 0.3)


def _check_format(fname: str, val) -> float:
    """Basic format plausibility check. Returns 0.0-1.0."""
    if not val: return 0.0
    s = str(val).strip()
    if not s: return 0.0
    # Date fields: should parse as date
    if any(kw in fname.lower() for kw in ("date", "dob", "expiry", "issued")):
        from datetime import datetime
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                datetime.strptime(s[:10], fmt)
                return 1.0
            except ValueError:
                continue
        return 0.3  # doesn't look like a date
    # Numeric fields: should contain numbers
    if any(kw in fname.lower() for kw in ("amount", "total", "price", "cost", "sum", "balance")):
        import re
        has_num = bool(re.search(r"\d", s))
        return 0.8 if has_num else 0.2
    # Email fields
    if "email" in fname.lower():
        return 0.9 if "@" in s else 0.1
    # Phone fields
    if any(kw in fname.lower() for kw in ("phone", "mobile", "tel", "contact")):
        import re
        return 0.8 if re.search(r"\d{3,}", s) else 0.2
    # Default: has reasonable length
    return 1.0 if len(s) >= 2 else 0.5


# ── G9 · multi-document split (detect-then-confirm) ──────────────────────

class SplitSegment(BaseModel):
    start_page: int
    end_page: int
    name: str | None = None


class SplitApplyPayload(BaseModel):
    segments: list[SplitSegment]


@router.post("/{doc_id}/split/suggest")
def suggest_document_split(
    doc_id: str,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """READ-ONLY: detect likely document boundaries inside a bundled PDF and
    return candidate segments. Suggests nothing destructive — the caller (UI)
    shows these, the user confirms, then calls /split/apply."""
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if not doc.s3_key:
        raise HTTPException(status_code=409, detail="Document has no stored file to analyze")
    blob = storage.get_object_bytes(doc.s3_key)
    if not blob:
        raise HTTPException(status_code=409, detail="Document file unavailable")
    from app.services import doc_split
    segments = doc_split.suggest_segments(blob)
    return {"docId": doc_id, "segments": segments,
            "splitSuggested": len(segments) > 1}


@router.post("/{doc_id}/split/apply", status_code=201)
async def apply_document_split(
    doc_id: str,
    payload: SplitApplyPayload,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Confirm-gated: split the parent PDF into one child Document per segment
    and enqueue ingestion for each. NON-DESTRUCTIVE — the parent is kept (archive
    it separately if desired)."""
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if not payload.segments:
        raise HTTPException(status_code=400, detail="No segments provided")
    from app.services import doc_split
    try:
        children = doc_split.apply_split(
            db, doc, [s.model_dump() for s in payload.segments], uploaded_by=user.email)
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))
    tenant = get_current_tenant()
    for c in children:
        try:
            await enqueue_ingest(c["pk"], tenant)
        except Exception as e:  # noqa: BLE001
            logging.getLogger("docaiq.routers.documents").warning(
                "split: enqueue_ingest failed for child pk=%s: %s", c["pk"], e)
    return {"parentId": doc_id, "created": [c["dict"] for c in children]}



class DocTypePayload(BaseModel):
    docType: str
    reason: str | None = None


@router.patch("/{doc_id}/type", response_model=Document)
def set_document_type(
    doc_id: str,
    payload: DocTypePayload,
    background: BackgroundTasks,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """HITL override of the document TYPE. The classifier picks from a fixed
    enum and can be wrong (or lack a matching type); this lets a reviewer set
    it by hand. Recorded as a field_edits row (field_path='doc_type') so it's
    audit-logged AND available to the P10 learning engine to bias future
    classification of similar docs. Sets confidence to 1.0 (human-verified)
    and re-fires the matcher (type affects requirement scoring)."""
    new_type = (payload.docType or "").strip()
    if not new_type:
        raise HTTPException(status_code=422, detail="docType is required")

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    from app.orm import FieldEdit
    original = doc.doc_type
    doc.doc_type = new_type
    doc.doc_type_confidence = 1.0          # human-verified
    doc.doc_type_alternatives = []
    db.add(FieldEdit(
        tenant_id=user.org_id,
        document_pk=doc.pk,
        field_path="doc_type",
        original_value=(original or "")[:4000],
        new_value=new_type[:4000],
        edited_by=user.email,
        reason=(payload.reason or "")[:2000] or None,
    ))
    # §2 · HITL feeds the self-learning vocabulary — register the human-set type
    # (top priority) + fold this doc into its centroid so similar future docs
    # distill straight to it. Best-effort; documents product only.
    try:
        from app.services.type_reconciler import learn_human_type
        learn_human_type(db, doc, new_type)
    except Exception:  # noqa: BLE001
        pass
    db.commit()
    import logging
    logging.getLogger("docaiq.routers.documents").info(
        "HITL doc_type override · doc=%s %r → %r by %s", doc_id, original, new_type, user.email,
    )
    # Type affects requirement matching — re-fire the matcher.
    background.add_task(enqueue_rematch, doc.pk, user.org_id)
    return repo.get(db, doc_id)


@router.get("/{doc_id}/edit-history")
def get_edit_history(
    doc_id: str,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(get_current_user),
) -> list[dict]:
    """Audit trail of every manual edit on this document, newest first."""
    from app.orm import FieldEdit
    from sqlalchemy import select as sa_select, desc

    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    tid = get_current_tenant()
    rows = db.scalars(
        sa_select(FieldEdit)
        .where(FieldEdit.tenant_id == tid, FieldEdit.document_pk == doc.pk)
        .order_by(desc(FieldEdit.edited_at))
    ).all()
    return [
        {
            "pk": r.pk,
            "fieldPath": r.field_path,
            "originalValue": r.original_value,
            "newValue": r.new_value,
            "editedBy": r.edited_by,
            "reason": r.reason,
            "editedAt": r.edited_at.isoformat() if r.edited_at else None,
        }
        for r in rows
    ]


@router.post("/{doc_id}/recategorize")
def recategorize_document(
    doc_id: str,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Re-run the categorizer over this doc's transactions / items. Useful
    after the merchant cache has been seeded or when a reviewer wants to
    refresh categories. Uses cache first so cost is usually $0."""
    import copy
    from app.agents import categorizer

    doc = repo.get_row(db, doc_id)
    if doc is None or not doc.extracted_fields:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} has no extracted fields")
    # Deep-copy so SQLAlchemy sees a new object reference and flushes the
    # JSONB column. In-place mutations of the existing dict don't dirty
    # the attribute (JSONB has no mutation tracking by default).
    ef = copy.deepcopy(doc.extracted_fields)
    fields = ef.get("fields") or {}
    total = 0
    cache_hits = 0
    for key in ("top_transactions", "items", "line_items"):
        arr = fields.get(key)
        if isinstance(arr, list) and arr:
            r = categorizer.categorize_transactions(
                db, user.org_id, arr, vendor_pk=doc.vendor_pk,
            )
            total += r.categorized
            cache_hits += r.cached_hits
    ef["fields"] = fields
    doc.extracted_fields = ef
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(doc, "extracted_fields")
    db.commit()

    # M28 · setting a real category may unblock auto-approve.
    try:
        from app.document_review import (
            get_document_threshold, get_duplicate_doc_ids, try_auto_approve,
        )
        threshold = get_document_threshold(db)
        if threshold is not None:
            if try_auto_approve(db, doc, threshold=threshold,
                                duplicate_doc_ids=get_duplicate_doc_ids(db)):
                db.commit()
    except Exception:  # noqa: BLE001
        pass

    return {"categorized": total, "cache_hits": cache_hits}


def _reapply_manual_field_edits(db: Session, doc) -> int:
    """After a fresh extraction OVERWRITES extracted_fields, re-apply the reviewer's manual
    top-level field edits / additions / deletions (from the FieldEdit audit) so human
    corrections are never silently lost on Re-extract. Human input = ground truth.
    Scope: top-level `fields.<key>` scalars only (nested array cells re-extract cleanly)."""
    import re as _re
    from app.orm import FieldEdit
    from sqlalchemy.orm.attributes import flag_modified
    edits = db.scalars(
        select(FieldEdit).where(FieldEdit.document_pk == doc.pk).order_by(FieldEdit.pk)
    ).all()
    if not edits:
        return 0
    # Collapse the audit log to the LATEST manual action per top-level field key.
    final: dict[str, tuple[str, str | None]] = {}
    for e in edits:
        m = _re.fullmatch(r"fields\.([a-z0-9_]+)", e.field_path or "")
        if not m:
            continue  # skip nested paths (fields.x.0.y) and non-field edits (__markdown__ etc.)
        key = m.group(1)
        if e.new_value is None and "delet" in (e.reason or "").lower():
            final[key] = ("del", None)
        elif e.new_value is not None:
            final[key] = ("set", e.new_value)
    if not final:
        return 0
    ef = doc.extracted_fields if isinstance(doc.extracted_fields, dict) else {}
    if not isinstance(ef.get("fields"), dict):
        ef["fields"] = {}
    fields = ef["fields"]
    changed = 0
    for key, (action, val) in final.items():
        cur = fields.get(key)
        if action == "del":
            if key in fields:
                fields.pop(key, None)
                for sub in ("field_confidence", "field_bboxes", "field_mentions"):
                    if isinstance(ef.get(sub), dict):
                        ef[sub].pop(key, None)
                changed += 1
        else:
            # Don't clobber a freshly-extracted ARRAY (transactions/line_items) with an old scalar.
            if isinstance(cur, (list, dict)):
                continue
            if cur != val:
                fields[key] = val
                ef.setdefault("field_confidence", {})[key] = 1.0  # human-reviewed
                changed += 1
    if changed:
        doc.extracted_fields = ef
        flag_modified(doc, "extracted_fields")
        db.commit()
    return changed


@router.post("/{doc_id}/reanalyze")
def reanalyze_document(
    doc_id: str,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Phase 3 · 'Re-analyze with the best model'. Re-runs the fact extractor on this doc
    using the STRONG model (config.strong_extract_model) while keeping its classification —
    for escalating a weak/thin extraction on demand. The reviewer then confirms the result;
    on sign-off it becomes a golden case + feeds crystallization (learn-then-downgrade)."""
    from app.rate_limit import rate_limit as _rate_limit
    _rate_limit(user.email, action="doc_reclassify")
    from app.agents import fact_extractor
    from app.config import get_settings

    row = repo.get(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if row.get("ingestionStatus") != "ready":
        raise HTTPException(status_code=409,
                            detail=f"Document is not ready (status={row.get('ingestionStatus')}).")
    doc_orm = repo.get_row(db, doc_id)
    if doc_orm is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    model = get_settings().strong_extract_model
    tok = fact_extractor._model_override.set(model)
    try:
        fx = fact_extractor.extract(db, document_pk=doc_orm.pk,
                                    classifier_doc_type=doc_orm.doc_type or "other")
    finally:
        fact_extractor._model_override.reset(tok)
    if fx is not None:
        doc_orm.extracted_fields = fx.to_jsonb()
        db.commit()
        try:
            from app.graph import bootstrap as graph_bootstrap
            graph_bootstrap.run(db, doc_orm.pk)
            db.commit()
        except Exception:  # noqa: BLE001
            db.rollback()
    # Re-apply the reviewer's manual field edits/additions so Re-analyze never wipes them.
    try:
        _reapply_manual_field_edits(db, doc_orm)
    except Exception:  # noqa: BLE001
        db.rollback()
    return {"model": model, "reanalyzed": fx is not None, "document": repo.get(db, doc_id)}


@router.post("/{doc_id}/reclassify", response_model=Document)
def reclassify_document(
    doc_id: str,
    db: Session = Depends(get_session),
    _user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    # M44.P9.12 · 10 reclassifies / min per user
    from app.rate_limit import rate_limit as _rate_limit
    _rate_limit(_user.email, action="doc_reclassify")
    """Re-run the classifier + fact-extractor on an already-ingested doc.

    Useful when (a) the original classification was wrong or low-confidence,
    (b) the doc was ingested before classifier/fact-extractor was wired up,
    (c) we've shipped a new schema and want to refresh older docs. Runs
    inline rather than via the worker queue so the caller gets the fresh
    classification in the response.

    Note: requires the doc to already be ingested (chunks exist) — image-only
    PDFs and failed uploads still need to come through the normal ingest path.
    """
    from app.agents import fact_extractor
    from app.agents.classifier import classify_document, persist as persist_classification

    row = repo.get(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if row.get("ingestionStatus") != "ready":
        raise HTTPException(
            status_code=409,
            detail=f"Document is not ready (status={row.get('ingestionStatus')}); re-upload it first.",
        )

    doc_orm = repo.get_row(db, doc_id)
    if doc_orm is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    # 1. Classifier
    result = classify_document(db, doc_orm.pk)
    if result is None:
        raise HTTPException(
            status_code=502,
            detail="Classifier returned no result — check OpenRouter key + worker logs.",
        )
    persist_classification(db, doc_orm.pk, result)

    # 2. Fact extractor (only fires for types mapped to a schema and conf ≥ 0.5)
    if result.top.confidence >= 0.5:
        fx = fact_extractor.extract(
            db,
            document_pk=doc_orm.pk,
            classifier_doc_type=result.top.doc_type,
        )
        if fx is not None:
            doc_orm.extracted_fields = fx.to_jsonb()
            db.commit()

    # 2b · M40 · Image documents — re-run the KYC vision extractor so the
    # reviewer-facing Re-extract button refreshes per-field bboxes (M40
    # Phase F adds RapidOCR + Tesseract OCR pass for deterministic field
    # positioning) without needing a delete + re-upload round-trip.
    # fact_extractor.extract() at step 2 is PDF-only and returns None for
    # images, so without this branch nothing happens for image uploads
    # (passport, driver's licence, ACRA bizfile screenshots).
    mime = (doc_orm.mime_type or "").lower()
    is_image = mime.startswith("image/")
    if is_image and doc_orm.s3_key:
        try:
            import logging
            from app.agents import kyc_extractor as kyc_x
            log_ = logging.getLogger("docaiq.routers.documents")
            # Resolution order for the KYC schema key:
            #   1. If the doc is attached to a KYC-* requirement, that
            #      requirement's mapped schema (KYC_REQUIREMENT_TO_DOC_TYPE).
            #      Picks the country-specific schema when the attachment
            #      pins it (passport_us, aadhaar, nric, …).
            #   2. Else fall back to the classifier's doc_type via
            #      CLASSIFIER_DOC_TYPE_TO_SCHEMA — generic schema
            #      ("primary_photo_id") for ad-hoc uploads not yet
            #      attached to any KYC-scoped requirement.
            from app.orm import Requirement
            from sqlalchemy import select
            tid = doc_orm.tenant_id
            attached_req_ids = list(db.scalars(
                select(Requirement.id_external).where(
                    Requirement.tenant_id == tid,
                    Requirement.doc_id_external == doc_orm.id_external,
                )
            ).all())
            kyc_doc_type = None
            for rid in attached_req_ids:
                kyc_doc_type = kyc_x.KYC_REQUIREMENT_TO_DOC_TYPE.get(rid)
                if kyc_doc_type:
                    log_.info(
                        "KYC re-extract: routing via requirement %s → schema %s",
                        rid, kyc_doc_type,
                    )
                    break
            if kyc_doc_type is None:
                # Map classifier's doc_type (e.g. "driver_licence") to a
                # schema key (e.g. "primary_photo_id"). This is the path
                # most ad-hoc uploads take.
                kyc_doc_type = kyc_x.CLASSIFIER_DOC_TYPE_TO_SCHEMA.get(result.top.doc_type)
                if kyc_doc_type:
                    log_.info(
                        "KYC re-extract: routing via classifier doc_type %s → schema %s",
                        result.top.doc_type, kyc_doc_type,
                    )
            if kyc_doc_type and kyc_doc_type in kyc_x.SCHEMAS:
                kyc_result = kyc_x.extract(
                    s3_key=doc_orm.s3_key,
                    mime=mime,
                    doc_type=kyc_doc_type,
                )
                if kyc_result is not None:
                    doc_orm.extracted_fields = kyc_x.result_to_jsonb(kyc_result)
                    db.commit()
                    log_.info(
                        "KYC re-extract: persisted doc pk=%s · %d fields · %d bboxes",
                        doc_orm.pk,
                        len(kyc_result.fields),
                        len(kyc_result.field_bboxes or {}),
                    )
                else:
                    log_.warning(
                        "KYC re-extract: kyc_extractor returned None for doc pk=%s (schema=%s, mime=%s)",
                        doc_orm.pk, kyc_doc_type, mime,
                    )
            else:
                log_.warning(
                    "KYC re-extract: no schema mapped for doc pk=%s · classifier_type=%s · attached_reqs=%s",
                    doc_orm.pk, result.top.doc_type, attached_req_ids,
                )
        except Exception as e:  # noqa: BLE001
            import logging
            logging.getLogger("docaiq.routers.documents").exception(
                "KYC re-extract failed for doc pk=%s: %s", doc_orm.pk, e,
            )

    # 3. Bootstrap the graph from the fresh facts.
    # Best-effort — failure here doesn't roll back the reclassify response.
    if doc_orm.extracted_fields:
        try:
            from app.graph import bootstrap as graph_bootstrap
            graph_bootstrap.run(db, doc_orm.pk)
            db.commit()
        except Exception as e:  # noqa: BLE001
            import logging
            logging.getLogger("docaiq.routers.documents").warning(
                "graph bootstrap failed for doc pk=%s: %s", doc_orm.pk, e,
            )

    # M28 · auto-approve check after reclassify.
    try:
        from app.document_review import (
            get_document_threshold, get_duplicate_doc_ids, try_auto_approve,
        )
        threshold = get_document_threshold(db)
        if threshold is not None:
            if try_auto_approve(db, doc_orm, threshold=threshold,
                                duplicate_doc_ids=get_duplicate_doc_ids(db)):
                db.commit()
    except Exception:  # noqa: BLE001
        pass

    # Re-apply the reviewer's manual field edits/additions so Re-extract never wipes them.
    try:
        _reapply_manual_field_edits(db, doc_orm)
    except Exception:  # noqa: BLE001
        db.rollback()

    fresh = repo.get(db, doc_id)
    if fresh is None:
        raise HTTPException(status_code=500, detail="Document disappeared after re-classification")
    return fresh


@router.delete("/{doc_id}")
def delete_document(
    doc_id: str,
    force: bool = Query(False, description="Reserved for future overrides; ignored today. Closed-audit refs always block hard delete."),
    db: Session = Depends(get_session),
    _user: CurrentUser = Depends(require_role("admin", "reviewer", "vendor")),
) -> dict:
    """Hard-delete a document while only ACTIVE audits reference it.

    Two-phase delete (M44.P10) when ``DOCAIQ_DELETE_WITH_LEARNING`` is on:
      · Phase 1 · learn-and-promote — lift generalizable knowledge (helpful
        reflexion pairs, repeated field corrections, successful agent skills,
        org/person canonicals) into the tenant UNDERSTANDING tables so it
        survives the cascade. Runs under a row lock; any failure rolls the
        whole request back (get_session) so the doc stays put.
      · Phase 2 · cascade delete — the existing sync-delete below.
    Flag-off, Phase 1 is skipped and behavior is exactly as before.

    Policy (M29 · 2026-05-23):
      · No-reference docs (dummy/mock uploads) → delete freely.
      · Referenced only by active audits → delete + cascade-clean orphan
        refs in requirements/highlights/chat/diffs/reflexion_pairs.
      · Referenced by any CLOSED audit → 409, with `closedAudits` in the
        detail. Caller should POST /archive instead to soft-hide and
        preserve history + next-cycle clones.

    Role: admin, reviewer, OR vendor. A vendor may delete only their OWN
    documents — `delete_row` → `get_row` carries the M17 vendor clause, so a
    vendor deleting another vendor's doc (or a tenant-shared doc) gets a 404,
    never a cross-vendor delete. The cascade (chunks, entities, pii_vault,
    artifacts via FK CASCADE + requirements/highlights/chat/diffs/reflexion
    cleanup) is the same complete sync-delete path for every role.

    Returns ``{id, learningPromoted}`` — `learningPromoted` is the Phase-1
    telemetry summary, or null when the flag is off.
    """
    # Ownership FIRST (M17): get_row carries the vendor clause, so a vendor
    # asking about a doc they don't own gets a clean 404 — before we reveal
    # anything about closed-audit references. Admin/reviewer (no vendor scope)
    # see every doc, so this only narrows vendor-role callers.
    pre = repo.get_row(db, doc_id)
    if pre is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    closed_ids = repo.referenced_by_closed_audit(db, doc_id)
    if closed_ids:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "doc_referenced_by_closed_audit",
                "message": (
                    f"Document {doc_id} is referenced by {len(closed_ids)} closed audit"
                    f"{'s' if len(closed_ids) != 1 else ''} — hard delete would break "
                    "history snapshots and the next-cycle clone. Archive instead."
                ),
                "closedAudits": closed_ids,
                "suggestedAction": "archive",
            },
        )

    # ── Phase 1 · learn-and-promote (flag-gated) ──────────────────────────
    promoted: dict | None = None
    if get_settings().delete_with_learning:
        import logging
        from app.services.learning_promoter import promote_doc_learnings
        try:
            promoted = promote_doc_learnings(db, pre.pk, lock=True)
        except Exception:  # noqa: BLE001
            # Roll back the whole request (get_session does this on raise) so
            # deletion_status reverts to NULL and the doc is NOT deleted.
            logging.getLogger("docaiq.routers.documents").exception(
                "Phase-1 learning promotion failed for %s; aborting delete", doc_id
            )
            raise HTTPException(
                status_code=500,
                detail={
                    "code": "learning_promotion_failed",
                    "message": "Learning promotion failed; document was NOT deleted. Retry.",
                },
            )
        else:
            logging.getLogger("docaiq.routers.documents").info(
                "delete-with-learning %s: %s", doc_id, promoted
            )

    # ── Phase 2 · cascade delete ──────────────────────────────────────────
    row = repo.delete_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if row.s3_key:
        try:
            storage.delete_object(row.s3_key)
        except Exception:
            pass   # S3 object missing is acceptable — DB row already gone.

    return {"id": doc_id, "learningPromoted": promoted}


@router.post("/{doc_id}/archive", response_model=Document)
def archive_document(
    doc_id: str,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Soft-archive a document. Idempotent. Use when hard-delete is refused
    (doc referenced by a closed audit) or when you just want to declutter
    the UI without losing history."""
    row = repo.archive_row(db, doc_id, by_email=user.email)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    return repo.get(db, doc_id)


@router.post("/{doc_id}/unarchive", response_model=Document)
def unarchive_document(
    doc_id: str,
    db: Session = Depends(get_session),
    _user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Restore a soft-archived document to the default list."""
    row = repo.unarchive_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    return repo.get(db, doc_id)


# ─────────────────────────────────────────────────────────────────────────
# M29.2 · per-doc Actions surface (admin+reviewer).
#
# When a doc shows "Reqs matched = 0" in AllDocuments, the reviewer needs
# a quick way to either delete it, re-fire the matcher, manually attach
# it to a specific requirement, or escalate via RFI. Delete + archive
# already exist; these two endpoints add the matcher + attach actions.
# ─────────────────────────────────────────────────────────────────────────
@router.post("/{doc_id}/rematch", status_code=202)
def rematch_document(
    doc_id: str,
    background: BackgroundTasks,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Re-fire the matcher cascade for one already-ingested document.

    The same enqueue_rematch path used after a review-status flip — but
    on-demand. Returns 202 immediately; results update via the standard
    document-status polling once the worker drains the job.
    """
    row = repo.get_row(db, doc_id)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")
    if row.ingestion_status != "ready":
        raise HTTPException(
            status_code=409,
            detail=f"Document {doc_id} is not ready (status={row.ingestion_status}); cannot re-match.",
        )
    background.add_task(enqueue_rematch, row.pk, user.org_id)
    return {"enqueued": True, "docId": doc_id}


class _AttachPayload(BaseModel):
    requirementId: str


@router.post("/{doc_id}/attach", response_model=Document)
def attach_document(
    doc_id: str,
    payload: _AttachPayload,
    db: Session = Depends(get_session),
    user: CurrentUser = Depends(require_role("admin", "reviewer")),
) -> dict:
    """Manually attach an unmatched document to a requirement.

    Differs from /api/requirements/{id}/attach-evidence (M28.7) — that one
    requires doc.review_status='reviewed' (reuse of already-vetted docs).
    This one accepts any ready doc, mirroring the wizard's HITL attach
    path for fresh uploads. Sets Requirement.doc_id_external = this doc,
    bumps status to 'warn' so the reviewer's verdict is still required.
    Confidence is left NULL — this is a human attach, not an AI match.
    """
    from sqlalchemy import select as sa_select
    from app.db import get_current_tenant
    from app.orm import Requirement as ReqORM

    tid = get_current_tenant()
    # M46 · owner-scoped lookup (no-op in auditing) so a documents-product user
    # can't attach another user's document by guessing its id (IDOR).
    doc = repo.get_row(db, doc_id)
    if doc is None:
        raise HTTPException(status_code=404, detail=f"Document {doc_id} not found")

    req = db.scalar(
        sa_select(ReqORM).where(ReqORM.tenant_id == tid, ReqORM.id_external == payload.requirementId)
    )
    if req is None:
        raise HTTPException(status_code=404, detail=f"Requirement {payload.requirementId} not found")

    # M31.6 · Multi-evidence. If primary already set, keep it as primary
    # but ALSO append this doc as additional evidence. If unset, this
    # becomes the primary AND first evidence entry.
    if not req.doc_id_external:
        req.doc_id_external = doc.id_external
        if req.status in ("todo", "miss"):
            req.status = "warn"
    from app.agents.matcher import _append_evidence
    _append_evidence(
        db, tid, req.pk, doc.id_external,
        confidence=None, source="manual", attached_by=user.email,
    )
    db.commit()
    return repo.get(db, doc_id)


# ── M53 · user annotations / highlights ──────────────────────────────────────
async def _doc_bytes(db: Session, row) -> bytes | None:
    """The document's original bytes — local blob, or re-pulled from Drive (M46
    self-heal) when autobackup purged the server copy (s3_key NULL, source=drive)."""
    if row.s3_key:
        return storage.get_object_bytes(row.s3_key)
    from app.config import get_settings as _gs
    if _gs().product == "documents" and row.source == "drive" and row.source_ref:
        from app.connectors import drive as drive_mod
        from app.repositories import connectors as conn_repo
        acct = conn_repo.get(db, "drive")
        if acct is not None:
            try:
                pulled = await drive_mod.get_backend().fetch(acct, row.source_ref)
                from app import drive_crypto
                return drive_crypto.decrypt_blob(row.owner_user_id, pulled.body)
            except Exception:  # noqa: BLE001
                return None
    return None


class AnnotationCreate(BaseModel):
    page: int = 1
    bbox: list[float]                 # [x0, y0, x1, y1] normalized 0..1
    note: str | None = None
    color: str | None = None


class AnnotationPatch(BaseModel):
    note: str | None = None
    color: str | None = None


@router.get("/{doc_id}/annotations/markdown")
def export_annotations_markdown(doc_id: str, db: Session = Depends(get_session),
                                user: CurrentUser = Depends(get_current_user)) -> dict:
    from app.repositories import annotations as arepo
    md = arepo.markdown_for_doc(db, doc_id)
    if md is None:
        raise HTTPException(status_code=404, detail="document not found")
    return {"markdown": md}


@router.get("/{doc_id}/annotations")
def list_annotations(doc_id: str, db: Session = Depends(get_session),
                     user: CurrentUser = Depends(get_current_user)) -> dict:
    from app.repositories import annotations as arepo
    res = arepo.list_for_doc(db, doc_id)
    if res is None:
        raise HTTPException(status_code=404, detail="document not found")
    return {"annotations": res}


@router.post("/{doc_id}/annotations", status_code=201)
async def create_annotation(doc_id: str, payload: AnnotationCreate,
                            db: Session = Depends(get_session),
                            user: CurrentUser = Depends(get_current_user)) -> dict:
    from app.repositories import annotations as arepo
    from app.services.region_capture import capture_region_text
    if not payload.bbox or len(payload.bbox) != 4:
        raise HTTPException(status_code=400, detail="bbox must be [x0,y0,x1,y1] normalized 0..1")
    d = arepo.resolve_doc(db, doc_id)
    if d is None:
        raise HTTPException(status_code=404, detail="document not found")
    x0, y0, x1, y1 = (float(v) for v in payload.bbox)
    body = await _doc_bytes(db, d)   # local blob, or Drive re-pull (purged docs)
    captured = ""
    if body:
        captured = capture_region_text(
            body, payload.page, x0, y0, x1, y1, db=db, tenant_id=get_current_tenant())
    res = arepo.create(db, doc_id, page=payload.page, x0=x0, y0=y0, x1=x1, y1=y1,
                       captured_text=captured, note=payload.note, color=payload.color)
    if res is None:
        raise HTTPException(status_code=404, detail="document not found")
    return res


@router.patch("/{doc_id}/annotations/{ann_id}")
def patch_annotation(doc_id: str, ann_id: int, payload: AnnotationPatch,
                     db: Session = Depends(get_session),
                     user: CurrentUser = Depends(get_current_user)) -> dict:
    from app.repositories import annotations as arepo
    res = arepo.update(db, ann_id, note=payload.note, color=payload.color)
    if res is None:
        raise HTTPException(status_code=404, detail="annotation not found")
    return res


@router.delete("/{doc_id}/annotations/{ann_id}")
def delete_annotation(doc_id: str, ann_id: int, db: Session = Depends(get_session),
                      user: CurrentUser = Depends(get_current_user)) -> dict:
    from app.repositories import annotations as arepo
    if not arepo.delete(db, ann_id):
        raise HTTPException(status_code=404, detail="annotation not found")
    return {"deleted": True}
